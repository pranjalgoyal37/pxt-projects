import re
from selenium import webdriver
from PIL import Image
import pytesseract
import cv2
import pandas as pd
import openpyxl as op
import time
import os
from os import listdir

excel_path=r"D:\Twitter _Scrrenshot_Taker\Excel_sheet\Infrigment.xlsx"
wb = op.load_workbook(excel_path)
sh1=wb['Sheet1']

sh2=wb['Sheet2']
print(sh1.max_row)
urls=[]
i=3

root_path = r"C:\Users\piexl\Downloads\Amazon_15-03-2023"

for folder in os.listdir(root_path):
    # print(folder)
    folder_path=os.path.join(root_path,folder)
    print(folder_path)
    
    for img_folder in os.listdir(folder_path):
        print(img_folder)
        img=os.path.join(folder_path,img_folder)
        
        img=cv2.imread(img)
        t=pytesseract.image_to_string(img)
        print(t)
        sh1.cell(i,column=2).value=str(t)
        i+=1
        wb.save(excel_path)

def read_data():
    df=pd.read_excel(r"D:\Twitter _Scrrenshot_Taker\Excel_sheet\Infringement.xlsx")
    urls=list(df['Url'])
    print(urls)


def img_read():
    img=cv2.imread("img.jpg")
    t=pytesseract.image_to_string(img)
    sh1.cell(i,column=2).value=str(t)
    i+=1
    wb.save(excel_path)

# read_data()
def take_ss():

    driver = webdriver.Chrome()
    for i in urls:
        driver.get(i)
        time.sleep(2)
        driver.save_screenshot("banner.png")
        img=cv2.imread("banner.jpg")
        t=pytesseract.image_to_string(img)
        sh1.cell(i,column=2).value=str(t)
        i+=1
        wb.save(excel_path)