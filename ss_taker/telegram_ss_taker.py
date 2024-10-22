import os 
from datetime import datetime
import xlsxwriter
import calendar
import openpyxl
import pandas as pd
# To take current date by using calender
date=datetime.now()
c_date=date.strftime('%d-%m-%y')
month =calendar.month_name[date.month]
platform='Telegram'


def create_folder():

    dir=r"D:\\"
    parent_folder='Amazon Associates '+platform

    ss_folder= platform+'_Sreenshots'
    sheet_folder=platform+'_Sheets'+"\\"

    country_name=['USA','Mexico','Canada','India']
    parent_path=os.path.join(dir,parent_folder)
    ss_path=os.path.join(dir,parent_folder,ss_folder)
    # month_path=os.path.join(dir,parent_folder,ss_folder,month)
    d_path=os.path.join(dir,parent_folder,ss_folder,c_date)

    os.chdir(dir)
    try:

        if not os.path.exists(parent_folder):
            os.mkdir(parent_folder)
            os.chdir(dir+parent_folder)
            os.mkdir(ss_folder)
            os.mkdir(sheet_folder)
            os.chdir(ss_path+"\\")
            os.mkdir(c_date)
            print(c_date+" date folder is created!",c_date)

            os.chdir(d_path+"\\")
            for country in country_name:
                if not os.path.exists(country):
                    os.mkdir(country)
                    print(country+" folder is created ")
                else:
                    print("folder is already Exsits:")

            
        
    # create screenshot folder
        elif not os.path.exists(parent_path+"\\"+ss_folder+"\\"):
            os.chdir(parent_path+"\\")
            os.mkdir(ss_folder)
            print("screenshot folder is created")

            os.chdir(ss_path+"\\")
            os.mkdir(c_date)
            print(c_date+" date folder is created!",c_date)

            os.chdir(d_path+"\\")
            for country in country_name:
                if not os.path.exists(country):
                    os.mkdir(country)
                    print(country+" folder is created ")
                else:
                    print("folder is already Exsits:")

     # create sheet folder
        elif not os.path.exists(parent_path+"\\"+sheet_folder+"\\"):
            os.chdir(parent_path+"\\")
            os.mkdir(sheet_folder)
            print("sheet folder is created")

        elif not os.path.exists(d_path+"\\"):
                print("If Part")
                os.chdir(ss_path+"\\")
                os.mkdir(c_date)
                print(c_date+" date folder is created!",c_date)
                os.chdir(d_path+"\\")
                for country in country_name:
                    if not os.path.exists(country):
                        os.mkdir(country)
                    else:
                        print("folder is already Exsits:")
        else:
            os.chdir(d_path+"\\")
            for country in country_name:
                    if not os.path.exists(country):
                        os.mkdir(country)
                    else:
                        print("folder is already Exsits:")  
    except OSError:
        print("folder is already exists")

def create_sheet():

    sheet_path=r"D:\Amazon Associates Telegram\Telegram_Sheets\\"
    #  Create a current date excel file 
    workbook = xlsxwriter.Workbook(sheet_path+'Amazon Associates Telegram '+str(c_date)+" .xlsx")
    worksheet = workbook.add_worksheet()
    sm_heading= ['id', 'channel_name', 'sub_channel_name', 'campaign_name', 'package_name', 'keyword_term', 'ad_description', 'ad_display_url', 'source_url', 'destination_url', 'destination_url_domain', 'publisher', 'sub_publisher', 'campaign', 'screenshot', ' inserted_date', 'category', 'status', 'type', 'coupon_code', 'brand', 'sm_handler', 'channel_id', 'upload_date', 'ad_heading', 'like', 'followers', 'members', 'contact', 'email_id', 'about_us_portal', 'priority', 'flag', 'customer', 'location', 'cessation', 'input_user', 'input_status', 'qc', 'qc_remarks', 'reviewed_by', 'approved_by', 'reviewed_comments', 'review ed_status', 'approvd_status', 'approved_comments', 'reviewed_date', 'approved_date', 'review_post_date', 'review_heading', 'ideation', 'user_type', 'review_like', 'review_followers', 'review_handler', 'review_rating', 'review_numbers', 'public_ response', 'review_comments', 'review_description', 'sub_sub_channel_name', 'country', 'sub_category', 'sub_sub_category', 'status_update_date', 'case_reports', 'intermediate_url']

    column = 0
    row=0
    for item in sm_heading :
        worksheet.write(row, column, item)
        column += 1
    print("Sheet is created ")
    workbook.close()
    
    if not os.path.exists(r"D:\Amazon Associates Telegram\\telegram_case.xlsx"):
       
        workbook = xlsxwriter.Workbook(r"D:\Amazon Associates Telegram\\telegram_case.xlsx")
        worksheet = workbook.add_worksheet()
        sm_heading=['ad_description', 'ad_display_url', 'source_url', 'destination_url']
        column=0
        for item in sm_heading :
            worksheet.write(0, column, item)
            column += 1
            print("sheet is created")
        workbook.close()
    else:
         print("Sheet is already created ")

# create_folder()
# create_sheet()

def read_data():
    df=pd.read_excel(r"D:\Amazon Associates Telegram\telegram_case.xlsx")


    ad_description=list(df['ad_description'])
    ad_display=list(df.ad_display_url)
    source_url=list(df['source_url'])
    destination_url=list(df['destination_url'])
    
    for data in df:
        print(data)

    # print(ad_description,ad_display,source_url,destination_url)
    # print("hello world")
    # for i in ad_description:
    #     print(i)
    # print(ad_description)
   
read_data()
   







